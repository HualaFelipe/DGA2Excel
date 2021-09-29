from DatosEstacion import DatosEstacion
from LectorCelda import LectorCelda
import xlrd
import os
import sys
from openpyxl import Workbook

sys.setrecursionlimit(1500)

carpeta = "precipitacion"

directory = os.fsencode("./"+carpeta)
books = []
print("inicio lectura de archivos")
for file in os.listdir(directory):

    filename = os.fsdecode(file)
    books.append(xlrd.open_workbook(carpeta+str("/"+filename)))

lectores = []
print("inicio lectura de hojas")
for book in books:

    for i in range(book.nsheets):
        lc = LectorCelda(book.sheet_by_index(i))
        lc.evaluarHoja()
        lectores.append(lc)

datosEstaciones = []

print("Estaciones")
for lector in lectores:

    ## Estacion existe ??
    existe = False
    index = 0
    for datosEstacion in datosEstaciones:
        if(datosEstacion.estacion == lector.hojaDatosDGA.estacion):
            existe = True
            break
        index = index+1
    ## Si existe cargala ahí
    if(existe):
        datosEstaciones[index].agregarHojaDatos(lector.hojaDatosDGA)
    ## Si no existe crea una nueva
    else:
        datosEstaciones.append(DatosEstacion(lector.hojaDatosDGA))


contEst = 0

print("Archivos")
for datosEstacion in datosEstaciones:
    wb = Workbook()
    ws = wb.active
    ws.title = "Serie de tiempo"

    ws['A1'] = "Tiempo"
    ws['B1'] = "Dato"
    i = 2;
    for dataTime in datosEstacion.serieTiempo:
        ws['A'+str(i)] = dataTime.dt
        ws['B' + str(i)] = dataTime.valor
        i = i+1


    wsGeneralData = wb.create_sheet("Datos Generales")

    wsGeneralData['A1'] = 'Estacion'
    wsGeneralData['A2'] = 'CodigoBNA'
    wsGeneralData['A3'] = 'Cuenca'
    wsGeneralData['A4'] = 'Subcuenca'
    wsGeneralData['A5'] = 'Altitud'
    wsGeneralData['A6'] = 'Latitud'
    wsGeneralData['A7'] = 'Longitud'
    wsGeneralData['A8'] = 'UTM Norte'
    wsGeneralData['A9'] = 'UTM Este'
    wsGeneralData['A10'] = 'Àrea Drenaje'


    wsGeneralData['B1'] = datosEstacion.estacion
    wsGeneralData['B2'] = datosEstacion.codigoBNA
    wsGeneralData['B3'] = datosEstacion.cuenca
    wsGeneralData['B4'] = datosEstacion.subcuenca
    wsGeneralData['B5'] = datosEstacion.altitud
    wsGeneralData['B6'] = datosEstacion.latitud
    wsGeneralData['B7'] = datosEstacion.longitud
    wsGeneralData['B8'] = datosEstacion.UTMNorte
    wsGeneralData['B9'] = datosEstacion.UTMEste
    wsGeneralData['B10'] = datosEstacion.areaDrenaje



    ##Guarar los datos generales

    wb.save("./out/"+str(contEst)+"_"+str(datosEstacion.estacion)+'.xlsx')

    print("./out/"+str(contEst)+"_"+str(datosEstacion.estacion)+'.xlsx')
    contEst = contEst +1



print("Listop!")
print("xd")
